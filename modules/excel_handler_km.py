import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

HEADER_ROW = 1
DATA_START_ROW = 2

PAYS_MAP = {
    "FR": "France", "BE": "Belgium", "DE": "Germany", "NL": "Netherlands",
    "LU": "Luxembourg", "IT": "Italy", "ES": "Spain", "PT": "Portugal",
    "GB": "United Kingdom", "CH": "Switzerland", "AT": "Austria",
    "PL": "Poland", "CZ": "Czech Republic", "HU": "Hungary",
    "RO": "Romania", "BG": "Bulgaria", "SK": "Slovakia", "SI": "Slovenia",
    "HR": "Croatia", "DK": "Denmark", "SE": "Sweden", "NO": "Norway",
    "FI": "Finland", "IE": "Ireland", "GR": "Greece",
    "F": "France", "B": "Belgium", "D": "Germany", "I": "Italy",
    "E": "Spain", "L": "Luxembourg", "A": "Austria", "P": "Portugal",
}

CP_LENGTHS = {
    "FR": 5, "F": 5,
    "BE": 4, "B": 4,
    "DE": 5, "D": 5,
    "IT": 5, "I": 5,
    "ES": 5, "E": 5,
    "PL": 5, "CZ": 5, "HR": 5, "SK": 5, "GR": 5, "SE": 5, "FI": 5,
    "NL": 4,
    "AT": 4, "A": 4,
    "CH": 4,
    "HU": 4, "DK": 4, "SI": 4, "NO": 4,
    "LU": 4, "L": 4,
    "PT": 7, "P": 7,
    "RO": 6,
}

CITY_CORRECTIONS = {
    "Basse Indre":    ("Indre", "44610", "FR"),
    "Saint Herblain": ("Saint-Herblain", "44800", "FR"),
    "Montataire":     ("Montataire", "60160", "FR"),
}

ZONE_CORRECTIONS = {
    # ITALIE
    "02000, Italy": "Rieti, 02100, Italy",
    "06000, Italy": "Perugia, 06100, Italy",
    "10000, Italy": "Torino, 10100, Italy",
    "16000, Italy": "Genova, 16100, Italy",
    "20000, Italy": "Milano, 20100, Italy",
    "21000, Italy": "Varese, 21100, Italy",
    "23000, Italy": "Sondrio, 23100, Italy",
    "24000, Italy": "Bergamo, 24100, Italy",
    "25000, Italy": "Brescia, 25100, Italy",
    "30000, Italy": "Mestre, 30170, Italy",
    "31000, Italy": "Treviso, 31100, Italy",
    "33000, Italy": "Udine, 33100, Italy",
    "36000, Italy": "Vicenza, 36100, Italy",
    "00000, Italy": "Roma, 00133, Italy",
    "50000, Italy": "Firenze, 50127, Italy",
    "80000, Italy": "Napoli, 80146, Italy",
    # PAYS-BAS
    "6000, Netherlands": "Weert, 6000, Netherlands",
    "6200, Netherlands": "Maastricht, 6221, Netherlands",
    "6600, Netherlands": "Wijchen, 6602, Netherlands",
    "1000, Netherlands": "Amsterdam, 1043, Netherlands",
    "3000, Netherlands": "Rotterdam, 3089, Netherlands",
    # BELGIQUE
    "8800, Belgium": "Roeselare, 8800, Belgium",
    "1000, Belgium": "Bruxelles, 1120, Belgium",
    "2000, Belgium": "Antwerpen, 2030, Belgium",
    "4000, Belgium": "Liège, 4020, Belgium",
    # ALLEMAGNE
    "50000, Germany": "Köln, 50769, Germany",
    "58000, Germany": "Hagen, 58099, Germany",
    "59000, Germany": "Hamm, 59067, Germany",
    "10000, Germany": "Berlin, 13405, Germany",
    "20000, Germany": "Hamburg, 21129, Germany",
    "80000, Germany": "München, 80939, Germany",
    "60000, Germany": "Frankfurt, 60549, Germany",
    # FRANCE
    "75000, France": "Paris, 75012, France",
    "69000, France": "Vénissieux, 69200, France",
    "13000, France": "Marseille, 13015, France",
    "33000, France": "Bordeaux, 33300, France",
    "59000, France": "Lille, 59160, France",
    "67000, France": "Strasbourg, 67100, France",
    # ESPAGNE
    "28000, Spain": "Madrid, 28052, Spain",
    "08000, Spain": "Barcelona, 08040, Spain",
    "46000, Spain": "Valencia, 46024, Spain",
    "01000, Spain": "Vitoria-Gasteiz, 01015, Spain",
    "03000, Spain": "Alicante, 03008, Spain",
    "06000, Spain": "Badajoz, 06006, Spain",
    "09000, Spain": "Burgos, 09007, Spain",
    "30000, Spain": "Murcia, 30169, Spain",
    "36000, Spain": "Pontevedra, 36158, Spain",
    # AUTRICHE
    "1000, Austria": "Wien, 1110, Austria",
    # LUXEMBOURG
    "3400, Luxembourg": "Dudelange, 3400, Luxembourg",
    "1000, Luxembourg": "Luxembourg, 1000, Luxembourg",
    # BULGARIE
    "15000, Bulgaria": "Sofia, 1528, Bulgaria",
}

# Mots-clés qui indiquent une ligne d'en-tête parasite
HEADER_KEYWORDS = {
    "depart", "départ", "origin", "pays", "country", "cp", "code postal",
    "postal code", "destination", "dest", "city", "ville", "zone",
    "country of destination", "dest cntry", "dest zone txt", "dest reg",
    "postal code destination", "city of destination",
}

# Corrections GPS pour villes ambiguës
GPS_FIXES_ORIGIN = {
    "rumbek":  "Rumbeke, 8800, Belgium",
    "rumbeke": "Rumbeke, 8800, Belgium",
    "indre":   "Indre, 44610, France",
}


# ============================================================
# UTILITAIRES
# ============================================================

def find_first_empty_column(ws):
    col = 1
    while col < 100:
        if ws.cell(row=HEADER_ROW, column=col).value is None:
            return col
        col += 1
    return col


def pad_postal_code(cp, country_prefix):
    if not cp:
        return ""
    target_len = CP_LENGTHS.get(country_prefix.upper(), 5)
    if target_len == 0:
        return cp
    if len(cp) < target_len:
        cp = cp + "0" * (target_len - len(cp))
    return cp


def is_header_row(origin_raw, country_raw, postal_raw, city_raw):
    """Retourne True si la ligne ressemble à une ligne d'en-tête parasite."""
    values = [str(v).strip().lower() for v in [origin_raw, country_raw, postal_raw, city_raw] if v]
    return any(v in HEADER_KEYWORDS for v in values)


# ============================================================
# PARSEURS
# ============================================================

def parse_origin(origin_str):
    if not origin_str:
        return ""
    origin_str = str(origin_str).strip()

    last_word = origin_str.split()[-1].lower() if origin_str.split() else ""
    if last_word in GPS_FIXES_ORIGIN:
        return GPS_FIXES_ORIGIN[last_word]

    if "(" in origin_str and ")" in origin_str:
        ville = origin_str[:origin_str.index("(")].strip()
        code  = origin_str[origin_str.index("(") + 1:origin_str.index(")")].strip()

        pays_prefix = ""
        cp = ""
        for i, c in enumerate(code):
            if c.isdigit():
                pays_prefix = code[:i]
                cp = code[i:]
                break

        cp   = pad_postal_code(cp, pays_prefix)
        pays = PAYS_MAP.get(pays_prefix.upper(), pays_prefix)

        if ville in CITY_CORRECTIONS:
            ville_corr, cp_corr, pays_corr = CITY_CORRECTIONS[ville]
            pays_corr_full = PAYS_MAP.get(pays_corr.upper(), pays_corr)
            return f"{ville_corr}, {cp_corr}, {pays_corr_full}"

        return f"{ville}, {cp}, {pays}"

    return origin_str


def parse_destination(city, postal_code, country):
    country     = str(country     or "").strip()
    city        = str(city        or "").strip()
    postal_code = str(postal_code or "").strip()

    cp_pays = ""
    cp_num  = ""
    if "-" in postal_code:
        parts   = postal_code.split("-", 1)
        cp_pays = parts[0].strip()
        cp_num  = parts[1].strip()
    else:
        cp_num = postal_code

    pays_full = PAYS_MAP.get(country.upper(), country)
    if not pays_full or pays_full == country:
        pays_full = PAYS_MAP.get(cp_pays.upper(), pays_full or cp_pays)

    pays_prefix = cp_pays if cp_pays else country
    cp_num = pad_postal_code(cp_num, pays_prefix)

    if city.lower() in ("all cities", "all", ""):
        base_zone = f"{cp_num}, {pays_full}"
        return ZONE_CORRECTIONS.get(base_zone, base_zone)
    else:
        return f"{city}, {cp_num}, {pays_full}"


# ============================================================
# LECTURE EXCEL
# ============================================================

def find_column_by_headers(ws, keywords, scan_rows=3, after_col=None):
    """
    Cherche la première colonne dont le header matche un keyword.
    after_col : si précisé, ignore les colonnes <= after_col
    """
    for row in range(1, scan_rows + 1):
        for col in range(1, ws.max_column + 1):
            if after_col and col <= after_col:
                continue
            val = ws.cell(row=row, column=col).value
            if val and str(val).strip() in keywords:
                return col
    return None

def parse_origin_from_parts(city, postal_code, country):
    """
    Parse l'origine depuis 3 colonnes séparées (PAYS / CP / LOCALITE).
    Exemple : city="Hooimeersstraat 8, Wielsbeke", cp="8710", pays="B"
    → "Hooimeersstraat 8 Wielsbeke 8710 Belgium"
    """
    country     = str(country     or "").strip()
    city        = str(city        or "").strip()
    postal_code = str(postal_code or "").strip()

    pays_full = PAYS_MAP.get(country.upper(), country)
    cp_num    = pad_postal_code(postal_code, country)

    if city:
        return f"{city}, {cp_num}, {pays_full}"
    elif cp_num:
        return f"{cp_num}, {pays_full}"
    else:
        return pays_full


def read_all_sheets(filepath):
    wb = openpyxl.load_workbook(filepath)
    sheets_data = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # ── Détection colonnes ORIGINE (3 colonnes séparées) ─────────────────
        col_orig_pays = find_column_by_headers(ws, [
            "PAYS", "Pays", "pays", "Country", "COUNTRY",
            "Origin country", "Pays départ", "Pays depart"
        ])

        col_orig_cp = find_column_by_headers(ws, [
            "CODE POSTAL", "Code postal", "code postal",
            "CP", "Postal Code", "POSTAL CODE", "CP départ", "CP depart"
        ], after_col=col_orig_pays) if col_orig_pays else None

        col_orig_city = find_column_by_headers(ws, [
            "LOCALITE", "Localité", "localite", "localité",
            "Origin", "Depart", "DEPART", "départ", "DÉPART",
            "Origine", "origine", "origin", "Ville départ", "Ville depart"
        ], after_col=col_orig_cp or col_orig_pays) if col_orig_pays else None

        # ── Détection colonnes DESTINATION ───────────────────────────────────
        dest_after = col_orig_city or col_orig_cp or col_orig_pays

        col_dest_pays = find_column_by_headers(ws, [
            "PAYS", "Pays", "pays", "Country", "COUNTRY",
            "Country of destination", "Dest cntry"
        ], after_col=dest_after)

        col_dest_cp = find_column_by_headers(ws, [
            "CP destination", "CP dech", "CP déchargement",
            "Postal code destination", "Dest reg",
            "CODE POSTAL", "Code postal", "code postal",
            "CP", "Postal Code", "POSTAL CODE"
        ], after_col=col_dest_pays or dest_after)

        col_dest_city = find_column_by_headers(ws, [
            "City of destination", "Dest zone txt",
            "VILLE", "Ville", "ville", "City", "CITY",
            "Destination", "destination"
        ], after_col=col_dest_cp or col_dest_pays or dest_after)

        # ── Validation minimale ───────────────────────────────────────────────
        if not col_orig_pays or not col_dest_pays:
            print(f"  ⚠️  Feuille '{sheet_name}' ignorée "
                  f"[orig_pays={col_orig_pays}, dest_pays={col_dest_pays}]")
            continue

        print(f"  🔍 Feuille '{sheet_name}' — colonnes détectées :")
        print(f"      ORIGINE  → pays={col_orig_pays}, cp={col_orig_cp}, city={col_orig_city}")
        print(f"      DEST     → pays={col_dest_pays}, cp={col_dest_cp}, city={col_dest_city}")

        routes       = []
        row          = DATA_START_ROW
        lignes_vides = 0

        while row <= ws.max_row + 5:

            # Lecture origine
            orig_pays_raw = ws.cell(row=row, column=col_orig_pays).value
            orig_cp_raw   = ws.cell(row=row, column=col_orig_cp).value   if col_orig_cp   else ""
            orig_city_raw = ws.cell(row=row, column=col_orig_city).value if col_orig_city else ""

            # Lecture destination
            dest_pays_raw = ws.cell(row=row, column=col_dest_pays).value
            dest_cp_raw   = ws.cell(row=row, column=col_dest_cp).value   if col_dest_cp   else ""
            dest_city_raw = ws.cell(row=row, column=col_dest_city).value if col_dest_city else ""

            # Ligne vide ?
            if not orig_pays_raw and not dest_pays_raw:
                lignes_vides += 1
                if lignes_vides >= 3:
                    break
                row += 1
                continue

            lignes_vides = 0

            # Ligne d'en-tête parasite ?
            if is_header_row(orig_pays_raw, dest_pays_raw, dest_cp_raw, dest_city_raw):
                print(f"  ⏭️  Ligne {row} ignorée (en-tête parasite)")
                row += 1
                continue

            # ── Parse origine : 3 colonnes séparées ──────────────────────────
            orig_cp_str   = str(orig_cp_raw   or "").strip()
            orig_city_str = str(orig_city_raw or "").strip()
            orig_pays_str = str(orig_pays_raw or "").strip()

            origin = parse_origin_from_parts(orig_city_str, orig_cp_str, orig_pays_str)

            # Debug : affiche ce qui est transmis à PTV
            print(f"      [Ligne {row}] Origine parsée : '{origin}'")

            # ── Parse destination ─────────────────────────────────────────────
            dest = parse_destination(dest_city_raw, dest_cp_raw, dest_pays_raw)

            label = (f"{orig_city_str or orig_pays_str} ({orig_cp_str}) "
                     f"→ {dest_city_raw or 'Zone'} ({dest_cp_raw})")

            routes.append({
                "row":    row,
                "origin": origin,
                "dest":   dest,
                "label":  label,
            })
            row += 1

        if routes:
            sheets_data[sheet_name] = (ws, routes)
            print(f"  ✅ Feuille '{sheet_name}' : {len(routes)} route(s) extraite(s)")
        else:
            print(f"  ⚠️  Feuille '{sheet_name}' : aucune route valide trouvée")

    return wb, sheets_data



# ============================================================
# ÉCRITURE RÉSULTATS
# ============================================================

def unmerge_and_write(ws, row, col, value, font=None, fill=None, alignment=None, border=None):
    """Écrit dans une cellule en défusionnant si nécessaire."""
    for merge_range in list(ws.merged_cells.ranges):
        if (merge_range.min_row <= row <= merge_range.max_row and
                merge_range.min_col <= col <= merge_range.max_col):
            ws.unmerge_cells(str(merge_range))
            break
    cell = ws.cell(row=row, column=col, value=value)
    if font:      cell.font      = font
    if fill:      cell.fill      = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border
    return cell


def find_safe_col(ws, header_row, start_col):
    """Trouve la première colonne libre (pas fusionnée, pas de valeur)."""
    col = start_col
    while True:
        is_merged = False
        for merge_range in ws.merged_cells.ranges:
            if (merge_range.min_row <= header_row <= merge_range.max_row and
                    merge_range.min_col <= col <= merge_range.max_col):
                col = merge_range.max_col + 1
                is_merged = True
                break
        if not is_merged:
            val = ws.cell(row=header_row, column=col).value
            if val is None:
                return col
            col += 1


def write_km_results(ws, results, calculer_peage=False):
    raw_start  = find_first_empty_column(ws)
    safe_start = find_safe_col(ws, HEADER_ROW, raw_start)

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center")
    border = Border(
        left=Side(style="thin"),  right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin"),
    )

    headers = ["KM PTV", "Carte PTV"]
    if calculer_peage:
        headers.append("Péage (€)")

    col_map = {}
    for i, h in enumerate(headers):
        target_col = find_safe_col(ws, HEADER_ROW, safe_start + i)
        unmerge_and_write(ws, HEADER_ROW, target_col, h,
                          font=header_font, fill=header_fill,
                          alignment=header_align, border=border)
        col_map[h] = target_col

    # Style lien hypertexte
    link_font = Font(color="0563C1", underline="single")

    for r in results:
        row  = r["row"]
        data = r.get("data")
        if not data:
            continue

        # ── KM ───────────────────────────────────────────────────────────────
        unmerge_and_write(ws, row, col_map["KM PTV"], data["km"], border=border)

        # ── CARTE : URL publique cliquable ────────────────────────────────────
        carte_url = data.get("carte_url", "")
        if carte_url:
            cell = unmerge_and_write(
                ws, row, col_map["Carte PTV"],
                "🗺️ Voir carte",
                font=link_font,
                border=border,
            )
            cell.hyperlink = carte_url      # ← https:// → s'ouvre dans le navigateur
            cell.style = "Hyperlink"
        else:
            unmerge_and_write(ws, row, col_map["Carte PTV"], "", border=border)

        # ── PÉAGE ─────────────────────────────────────────────────────────────
        if calculer_peage:
            unmerge_and_write(
                ws, row, col_map["Péage (€)"],
                data.get("prix_peage", 0.0),
                border=border,
            )
